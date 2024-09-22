import asyncio
import json
import os.path
import traceback
from re import sub as re_sub
from typing import Union

import openpyxl
from openpyxl.styles import Alignment
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from pcrapi.bsdk.model import AccInfo
from pcrapi.bsdk.sdkclients import bsdkclient
from data_db import nk, config_db, result_db, acc_db, TimeUtils
from pcrapi.core import pcrclient
from pcrapi.game.db import database, dbstart
from pcrapi.game.db.dbmgr import instance as dbmgr
from pcrapi.game.model.requests import LoadIndexRequest


class GameData(object):
    _inited = False

    def __init__(self):
        self._love_story_dict = None
        self._all_units_simple_dict = None

    def db_initializer(self):
        if not self._inited:
            asyncio.new_event_loop().run_until_complete(dbstart.db_start())
            self._inited = True

    @property
    def all_units_simple_dict(self):
        if not self._all_units_simple_dict:
            temp_dict = {}
            for k, v in database.db.unlock_unit_dict.items():
                temp_dict[k] = v.unit_name
            self._all_units_simple_dict = temp_dict
        return self._all_units_simple_dict

    @property
    def all_units_dict(self):
        return database.db.unlock_unit_dict

    @property
    def unit_data_dict(self):
        return database.db.unit_data

    @property
    def pure_memory_to_unit_dict(self):
        return database.db.pure_memory_to_unit

    @property
    def unit_unique_equip_dict(self):
        return database.db.unit_unique_equip

    @property
    def love_story_dict(self):
        if self._love_story_dict is None:
            data = database.db.unit_story
            temp_dict = {}
            for item in data:
                temp_dict.setdefault(item.story_group_id, []).append(item.story_id)
            self._love_story_dict = temp_dict
        return self._love_story_dict


gamedata = GameData()


class PcrAccount(object):
    def __init__(self, username: str, password: str):
        # self._game_data = gamedata
        # self._user_info = None
        # self._user_units_data = None
        self._time = ''
        self._user_json = ''
        self._account_data = {'username': username, 'password': password}
        self._client = self._get_client()

    def _get_client(self):
        # Get Android Client
        sdkclient = bsdkclient(AccInfo(
            acc=self._account_data['username'],
            pwd=self._account_data['password']
        ))
        return pcrclient(sdkclient)

    async def login(self):
        try:
            await self._client.login()
            await self._request_json()
            # self._time = datetime.now().strftime('%Y-%m-%d_%H:%M:%S')
            self._time = TimeUtils.stamp_utcnow()
        except Exception as _e:
            traceback.print_exc()
            return -1, str(_e)
        else:
            return 0, ''

    async def _request_json(self):
        req = LoadIndexRequest()
        req.carrier = 'CR_NMSL'
        self._user_json = (await self._client.request(req)).json()
        pass

    @property
    def sum(self):
        """
        数据集合,time为POSIX的UTC时间戳（2038警告
        """
        return {'acc': self._account_data['username'], 'time': self._time, 'json': self._user_json}


class JsonExporter(object):
    def __init__(self):
        pass

    def export_json(self, path: str):
        # 匹配非法字符的正则表达式
        __pattern = r"[\/\\\:\*\?\"\<\>\|\u0000-\u001F\u007F-\u009F]"
        __results = result_db.db.all()
        if __results:
            try:
                for item in __results:
                    __query = acc_db.get_query()
                    filename = acc_db.db.search(__query['acc'] == item['acc'])[0]['alias']
                    filename = item['acc'] if not filename else filename
                    filename = re_sub(
                        pattern=__pattern,
                        string=f"{filename}_{TimeUtils.obj_localtime(item['time']).strftime(r'%Y-%m-%d_%H%M%S')}.json",
                        repl='')

                    with open(os.path.join(path, filename), mode='w') as f:
                        f.write(item['json'])
            except Exception as e:
                raise e


class ExcelExporter(object):
    def __init__(self, mode: bool = False):
        """
        导出Excel表格
        :param mode: bool, True代表分离，False代表不分离，默认False
        """
        # self._data: Dict[int, dict] = copy.deepcopy(data)
        self._pre_processed = False
        self._data = None
        self._game_data = gamedata
        # self._preprocess_data()
        self._xlsx: Workbook = Workbook()
        self._mode = mode
        # self._init_template()

    def _init_template(self):
        _headers = [
            ('index', '序号'),
            ('name', '昵称/用户名'),
        ]
        _info_headers = [
            ('id', 'uid'),
            ('time', '更新时间'),
            ('mana', 'mana数'),
            ('jewel', '钻石数'),
            ('sow_stone', '母猪石数'),  # 母猪石
            ('star_cup', '星球杯数'),  # 星球杯
            ('all_hearts', '整心+心碎数'),  # 整心+心碎
            ('master_coin', '大师币数'),  # 大师币
        ]

        _unit_headers = [('status', '状态'), ('level', '等级'), ('rank', 'RANK'), ('equip', '装备'),
                         ('ub_level', 'UB'), ('sk1_level', 'sk1'), ('sk2_level', 'sk2'),
                         ('skex_level', 'skex'), ('unique_equip', '专武'), ('love_story', '好感等级'),
                         ('memory', '记忆碎片'),
                         # ('pure_memory', '纯净碎片')
                         ]
        # _nicknames = global_var.NICKNAMES

        self._0_name = '账号信息'
        self._0_info_index = {}

        self._1_name = 'BOX信息'
        self._1_info_index = {}
        self._1_data_index = {}
        # self._1_content_row = 3
        # self._selected_units = global_var.SELECTEDUNITS
        self._selected_units = config_db.selected_units_db.query('units',
                                                                 set(self._game_data.all_units_simple_dict.keys()))
        wb: Workbook = Workbook()
        wb.remove(wb['Sheet'])
        _alignment = Alignment(vertical='center')

        def _common_sheet(_ws: Worksheet, headers: list[tuple[str, str]], cnt_index: dict[str, int] = None,
                          sep_mode: bool = False, header_row: int = 2):
            if not sep_mode:
                assert cnt_index is not None
            for i, item in enumerate(headers, start=1):
                col = _ws.max_column
                if col == 1 and not any([cell.value for cell in _ws['A']]):
                    pass
                else:
                    col += 1

                _ws.cell(row=1, column=col, value=item[1])
                _ws.merge_cells(start_row=1, end_row=header_row, start_column=col, end_column=col)
                if not sep_mode:
                    cnt_index[item[0]] = col
                else:
                    _ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 5
            # def _info_sheet(_ws: Worksheet, cnt_index: dict, header_row: int = 2):
            #     # ws: Worksheet = wb.create_sheet(title=self._1_name, index=0)
            #     for i, item in enumerate(_info_headers, start=1):
            #         col = _ws.max_column + i
            #         _ws.cell(row=1, column=col, value=item[1])
            #         cnt_index[item[0]] = col
            #         _ws.merge_cells(start_row=1, end_row=header_row, start_column=col, end_column=col)

            # if mode:
            #     _cnt = 0
            #     # ColumnDimension(ws, bestFit=True, auto_size=True)
            #     # ws = wb.active
            #     for i in range(1, len(_info_headers) + 2):
            #         if _cnt == 2:
            #             break
            #         if i == 1:
            #             _ws.cell(row=1, column=i, value=f'序号')
            #             self._1_info_index['index'] = i
            #             _ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 5
            #         elif _info_headers[i - 2][0] == 'alias':
            #             col = _ws.max_column + 1
            #             _ws.cell(row=1, column=col, value=f'{_info_headers[i - 2][1]}')
            #             self._1_info_index[_info_headers[i - 2][0]] = col
            #         else:
            #             continue
            #         _cnt += 1
            #         _ws.merge_cells(start_row=1, end_row=2, start_column=i, end_column=i)
            #         _ws.cell(row=1, column=i).alignment = _alignment
            # else:
            #     for i in range(1, len(_info_headers) + 2):
            #         if i == 1:
            #             _ws.cell(row=1, column=i, value=f'序号')
            #             self._0_info_index['index'] = i
            #             _ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 5
            #         else:
            #             _ws.cell(row=1, column=i, value=f'{_info_headers[i - 2][1]}')
            #             self._0_info_index[_info_headers[i - 2][0]] = i
            #             _ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 10
            #         _ws.merge_cells(start_row=1, end_row=2, start_column=i, end_column=i)
            #         _ws.cell(row=1, column=i).alignment = _alignment

        def _units_sheet(_ws: Worksheet, headers: list[tuple[str, str]], cnt_index: dict[int, dict[str, int]]):
            _nicknames = nk.nicknames
            _common_sheet(_ws, [('', '角色')], sep_mode=True)

            for i, item1 in enumerate(self._selected_units, start=1):
                col1 = _ws.max_column
                if col1 == 1 and not any([cell.value for cell in _ws['A']]):
                    pass
                elif i == 1:
                    col1 += 1
                else:
                    col1 += 2
                _ws.merge_cells(start_row=1, end_row=1, start_column=col1 + 1, end_column=col1 + len(_unit_headers) - 1)
                # 填入名称
                unit_name = self._game_data.all_units_simple_dict[item1]
                if _nicknames:
                    if item1 in _nicknames.keys():
                        _ws.cell(row=1, column=col1 + 1,
                                 value="-".join([str(item1), str(_nicknames[item1]), unit_name]))
                    else:
                        _ws.cell(row=1, column=col1 + 1, value="-".join([str(item1), unit_name]))
                else:
                    _ws.cell(row=1, column=col1 + 1, value="-".join([str(item1), unit_name]))
                # 建立索引
                cnt_index[item1] = {}
                for j, item2 in enumerate(headers):
                    col2 = col1 + j
                    _ws.cell(row=2, column=col2, value=str(item2[1]))
                    _ws.column_dimensions[openpyxl.utils.get_column_letter(col2)].width = 11.5
                    cnt_index[item1][item2[0]] = col2

        if self._mode:
            ws = wb.create_sheet(title=self._0_name, index=1)
            _common_sheet(ws, _headers, self._0_info_index)
            _common_sheet(ws, _info_headers, self._0_info_index)

            ws = wb.create_sheet(title=self._1_name, index=2)
            _common_sheet(ws, _headers, self._1_info_index)
            _units_sheet(ws, _unit_headers, self._1_data_index)
        else:
            ws = wb.create_sheet(title=self._1_name, index=2)
            _common_sheet(ws, _headers, self._1_info_index)
            _common_sheet(ws, _info_headers, self._1_info_index)
            _units_sheet(ws, _unit_headers, self._1_data_index)

        self._xlsx = wb

    def preprocess_data(self, force: bool = False):
        def search_dict_list(k, v, dict_list: list[dict], default=None) -> dict:
            if default is None:
                default = {}
            rst = list(filter(lambda d: d.get(k) == v, dict_list))
            return rst[0] if rst else default

        if force or (not self._pre_processed):
            self._data = {}
            _data = result_db.db.all()
            mem_list = [15, 30, 100, 120, 150, 50]
            for i, user_sum in enumerate(_data):
                if 'pdata' not in user_sum or user_sum['pdata'].get('db_ver', '') != f'{dbmgr.ver}':
                    raw_dict = json.loads(user_sum['json'])
                    raw_info = raw_dict['user_info']
                    acc_q = acc_db.get_query()
                    q_alias = acc_db.db.search(acc_q['acc'] == user_sum['acc'])
                    q_alias = q_alias[0]['alias'] if q_alias else q_alias
                    user_info = {
                        'time': TimeUtils.obj_localtime(user_sum['time']).strftime(TimeUtils.format1),
                        # 'time': user_sum['time'],
                        'acc': user_sum['acc'],
                        'name': q_alias if q_alias else raw_info['user_name'],
                        'id': str(raw_info['viewer_id']),
                        'mana': sum(raw_dict['user_gold'].values()),
                        'jewel': sum([raw_dict['user_jewel']['free_jewel'], raw_dict['user_jewel']['paid_jewel']]),
                        'sow_stone': search_dict_list('id', 90005, raw_dict['item_list']).get('stock', 0),  # 母猪石
                        'star_cup': search_dict_list('id', 25001, raw_dict['item_list']).get('stock', 0),  # 星球杯
                        # 'heart_debris': search_dict_list('id', 140001, raw_dict['item_list']).get('stock', 0),  # 心碎
                        # 'heart': search_dict_list('id', 140000, raw_dict['item_list']).get('stock', 0),  # 公主心
                        'all_hearts': f"{search_dict_list('id', 140000, raw_dict['user_equip']).get('stock', 0)}+"
                                      f"{search_dict_list('id', 140001, raw_dict['user_equip']).get('stock', 0)}",
                        'master_coin': search_dict_list('id', 90008, raw_dict['item_list']).get('stock', 0),  # 大师币
                    }

                    user_units = {}
                    for unit in raw_dict['unit_list']:
                        key: int = unit['id']
                        # _new_data = {}
                        tmp_dict = {
                            'is_owned': True,
                            'status': unit['unit_rarity'],
                            'rank': unit['promotion_level'],
                            'level': unit['unit_level'],
                            'ub_level': unit['union_burst'][0]['skill_level'],
                            'sk1_level': 0 if len(unit['main_skill']) < 1 else
                            unit['main_skill'][0]['skill_level'],
                            'sk2_level': 0 if len(unit['main_skill']) < 2 else
                            unit['main_skill'][1]['skill_level'],
                            'skex_level': 0 if len(unit['ex_skill']) < 1 else
                            unit['ex_skill'][0]['skill_level'],
                            'equip': [],
                            'unique_equip': [],
                            'love_story': len(
                                set(self._game_data.love_story_dict[key // 100]) & set(raw_dict['read_story_ids'])),
                            'memory': search_dict_list('id', 30000 + (key // 100), raw_dict['item_list']).get('stock',
                                                                                                              0),
                            'pure_memory': search_dict_list('id', 31000 + (key // 100), raw_dict['item_list']).get(
                                'stock',
                                0)
                            if key in self._game_data.pure_memory_to_unit_dict.values() else '未实装'
                        }

                        tmp_list1 = []
                        tmp_list2 = []
                        for equip in unit['equip_slot']:
                            if equip['id'] == 999999 or not equip['is_slot']:
                                tmp_list1.append('-')
                            elif equip['is_slot']:
                                tmp_list1.append(str(equip['enhancement_level']))
                        tmp_dict['equip'] = "/".join(tmp_list1)

                        if unit['unique_equip_slot']:
                            for u_equip in unit['unique_equip_slot']:
                                tmp_list2.append(str(u_equip['enhancement_level']))
                        else:
                            tmp_list2.append('未实装')
                        tmp_dict['unique_equip'] = "/".join(tmp_list2)

                        user_units[key] = tmp_dict

                    for unit_key in (set(self._game_data.all_units_simple_dict.keys()) - set(user_units.keys())):
                        tmp_dict = {
                            'is_owned': False,
                            'status': '',
                            'star': self._game_data.unit_data_dict[unit_key].rarity,
                            'memory': search_dict_list('id', 30000 + (unit_key // 100),
                                                       raw_dict['item_list']).get('stock', 0),
                            'pure_memory': search_dict_list('id', 31000 + (unit_key // 100),
                                                            raw_dict['item_list']).get('stock', 0)
                            if unit_key in self._game_data.pure_memory_to_unit_dict.values() else '未实装',
                        }
                        if tmp_dict['memory'] >= self._game_data.all_units_dict[unit_key].count_1:
                            tmp_status = '可拼'
                        else:
                            tmp_status = '无'
                        tmp_dict['status'] = tmp_status
                        user_units[unit_key] = tmp_dict

                    self._data[user_sum['acc']] = {'user_info': user_info, 'user_units': user_units,
                                                   'db_ver': f'{dbmgr.ver}'}
                    rst_q = result_db.get_query()
                    result_db.db.upsert({'pdata': self._data[user_sum['acc']]},
                                        rst_q['acc'] == user_sum['acc'])
                else:
                    self._data[user_sum['acc']] = user_sum['pdata']
            # for user in self._data.values():
            #     # tmp_dict1: dict = user['user_info'].copy()
            #     # tmp_dict2: dict = user['user_units_data'].copy()
            #     #
            #     #
            #     # tmp_dict1['alias'] = tmp_dict1['alias'] if tmp_dict1['alias'] else tmp_dict1['name']
            #     # tmp_dict1.pop('name')
            #     #
            #     # tmp_dict1['all_hearts'] = str(tmp_dict1['heart']) + '+' + str(tmp_dict1['heart_debris'])
            #     # tmp_dict1.pop('heart')
            #     # tmp_dict1.pop('heart_debris')
            #     #
            #     # tmp_dict1['time'] = user['time']
            #     # user.pop('time')
            #     # 角色部分
            #     for key, unit in tmp_dict2.items():
            #         if unit['is_owned']:
            #             tmp_status = f"{unit['star']}星"
            #             # if unit['star'] < 6:
            #             #     # mem_list = [15, 30, 100, 120, 150, 50]
            #             #     if unit['pure_memory'] is not None:
            #             #         if unit['pure_memory'] >= 50 and tmp_dict1['star_cup'] >= 100 and unit['memory'] >= sum(
            #             #                 mem_list[unit['star']:]):
            #             #             tmp_status = f"{unit['star']}可6"
            #             #     elif unit['star'] < 5 and unit['memory'] >= mem_list[unit['star']]:
            #             #         for i in range(unit['star'] + 2, len(mem_list)):
            #             #             if unit['memory'] < sum(mem_list[unit['star']:i]):
            #             #                 tmp_status = f"{unit['star']}可{i - 1}"
            #             #                 break
            #             #         else:
            #             #             tmp_status = f"{unit['star']}可{len(mem_list) - 1}"
            #
            #             tmp_list1 = []
            #             tmp_list2 = []
            #             for equip in unit['equip']:
            #                 if equip is None:
            #                     tmp_list1.append('-')
            #                 else:
            #                     tmp_list1.append(str(equip))
            #             unit['equip'] = "/".join(tmp_list1)
            #             for equip in unit['unique_equip']:
            #                 if equip is None:
            #                     tmp_list2.append('未实装')
            #                 else:
            #                     tmp_list2.append(str(equip))
            #             unit['unique_equip'] = "/".join(tmp_list2)
            #         elif unit['memory'] >= self._GameData.all_units_dict[key].count_1:
            #             tmp_status = '可拼'
            #
            #             # if unit['pure_memory'] is not None:
            #             #     if unit['pure_memory'] >= 50 and tmp_dict1['star_cup'] >= 100 and unit['memory'] >= sum(
            #             #             mem_list[unit['star']:]) + self._GameData.all_units_dict[key].count_1:
            #             #         tmp_status = "可拼6"
            #             # elif unit['memory'] >= self._GameData.all_units_dict[key].count_1:
            #             #     tmp_status = f"可拼{unit['star']}"
            #             #     if unit['memory'] - self._GameData.all_units_dict[key].count_1 >= mem_list[unit['star']]:
            #             #         for i in range(unit['star'] + 2, len(mem_list)):
            #             #             if unit['memory'] - self._GameData.all_units_dict[key].count_1 < sum(
            #             #                     mem_list[unit['star']:i]):
            #             #                 tmp_status = f"可拼{i - 1}"
            #             #                 break
            #             #         else:
            #             #             tmp_status = f"可拼{len(mem_list) - 1}"
            #         else:
            #             tmp_status = '无'
            #         unit.pop('star')
            #         unit['status'] = tmp_status
            #
            #         if unit['pure_memory'] is None:
            #             unit['pure_memory'] = '未实装'
            #
            #     user['user_info'] = tmp_dict1
            #     user['user_units_data'] = tmp_dict2

            # self._data = sorted(self._data)
            self._pre_processed = True
        else:
            pass

    def _fill_content(self):
        wb = self._xlsx

        def _fill(_ws: Worksheet, data_key: str, cnt_index: Union[dict[str, int], dict[int, dict[str, int]]],
                  mode: int = 0, start_row: int = 2):
            """
            fill content
            :param _ws: Worksheet
            :param cnt_index: cnt_index
            :param mode: 0:common, 1:units; default:0
            :return:
            """
            row = start_row - 1
            for i, (acc, user_data) in enumerate(self._data.items(), start=1):
                row += 1
                if mode == 0:  # common
                    for k, c in cnt_index.items():
                        if k == 'index':
                            _ws.cell(row=row, column=c, value=i)
                        else:
                            _ws.cell(row=row, column=c, value=user_data[data_key].get(k, ''))
                elif mode == 1:  # units
                    for unit_id, items in cnt_index.items():
                        for k, c in items.items():
                            _ws.cell(row=row, column=c, value=user_data[data_key][unit_id].get(k, ''))

        # def old_fill(_ws: Worksheet, mode: bool, short_info: bool = False):
        #     # _ws: Worksheet = wb[self._1_name]
        #     row = _ws.max_row
        #
        #     # tmp_list = sorted(self._data)
        #     if mode:
        #         # _info_index = self._0_info_index if info_only else self._1_info_index
        #         for i in sorted(self._data):
        #             user = self._data[i]
        #             row += 1
        #             _ws.cell(row=row, column=1, value=i + 1)
        #             for key, item in user['user_info'].items():
        #                 if self._0_info_index.get(key, None):
        #                     _ws.cell(row=row, column=self._0_info_index[key], value=item)
        #                 else:
        #                     continue
        #     else:
        #         _info_index = self._1_info_index if short_info else self._0_info_index
        #         for i in sorted(self._data):
        #             user = self._data[i]
        #             row += 1
        #             _ws.cell(row=row, column=1, value=i + 1)
        #             for key, item in user['user_info'].items():
        #                 if _info_index.get(key, None):
        #                     _ws.cell(row=row, column=_info_index[key], value=item)
        #                 else:
        #                     continue
        #             for key, unit in user['user_units_data'].items():
        #                 if key not in self._1_data_index.keys():
        #                     continue
        #                 # if unit['is_owned'] is False:
        #                 #     for col in self._1_data_index[key].values():
        #                 #         ws.cell(row=row, column=col, value='未拥有')
        #                 # else:
        #                 #     # tmp_dict = {}
        #                 for item in unit:
        #                     if self._1_data_index[key].get(item, None):
        #                         _ws.cell(row=row, column=self._1_data_index[key][item], value=unit[item])
        #                     else:
        #                         continue
        #     # ColumnDimension(ws, bestFit=True)

        if self._mode:  # 分离
            # ws1: Worksheet = wb[self._0_name]
            # _fill(wb[self._0_name], True)
            # _fill(wb[self._1_name], False, short_info=True)
            _fill(wb[self._0_name], 'user_info', self._0_info_index, start_row=3)

        _fill(wb[self._1_name], 'user_info', self._1_info_index, start_row=3)
        _fill(wb[self._1_name], 'user_units', self._1_data_index, mode=1, start_row=3)

        self._xlsx = wb

    def export_xlsx(self, path='export.xlsx'):
        self._xlsx.save(path)
        self._init_template()
        self.preprocess_data()
        self._fill_content()
        self._xlsx.save(path)


if __name__ == "__main__":
    pass
